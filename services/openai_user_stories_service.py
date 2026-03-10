import os
import json
import requests
import base64
import time
from datetime import datetime
from typing import Dict, List
from copy import deepcopy
from openai import OpenAI
from dotenv import load_dotenv
from docx import Document
from docx.table import Table
from docx.shared import Inches, Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.oxml import OxmlElement
from docx.oxml.ns import qn
from PIL import Image
import pandas as pd
from io import BytesIO
from llama_parse import LlamaParse
from llama_index.llms.openai import OpenAI as LlamaOpenAI
from llama_index.embeddings.openai import OpenAIEmbedding
from llama_index.core import VectorStoreIndex, load_index_from_storage, StorageContext, Settings
from llama_index.core.node_parser import MarkdownElementNodeParser
from openpyxl.utils import get_column_letter

load_dotenv()


class OpenAIUserStoriesService:
    """
    Enterprise-grade User Stories generator.
    Pipeline:
    BRD Documents → LlamaIndex RAG → GPT-5.1 (JSON) → Excel Export
    """

    def __init__(
        self,
        model: str = "gpt-5.1",
        max_output_tokens: int = 14000,
        temperature: float = 0.2,
        timeout: int = 600
    ):
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError("OPENAI_API_KEY not configured")

        self.client = OpenAI(api_key=api_key)
        self.model = model
        self.max_output_tokens = max_output_tokens
        self.temperature = temperature
        self.timeout = timeout
        
        # Initialize LlamaIndex components for RAG
        self.llm = LlamaOpenAI(model="gpt-5.1", api_key=api_key)
        self.embed_model = OpenAIEmbedding(model="text-embedding-3-large", api_key=api_key)
        
        # Set global LlamaIndex settings
        Settings.embed_model = self.embed_model
        Settings.llm = self.llm
        
        # Persistent index for optimization
        self.index = None
        self.current_embeddings_path = None

    # ------------------------------------------------------------------
    # GPT RESPONSE HANDLING
    # ------------------------------------------------------------------

    def _extract_text(self, response) -> str:
        """
        Extract text from OpenAI Responses API response object.
        Supports both .output_text and structured .output[].content[] shapes.
        """
        if hasattr(response, "output_text") and response.output_text:
            return response.output_text.strip()

        texts = []
        for item in getattr(response, "output", []):
            for content in getattr(item, "content", []):
                if content.get("type") == "output_text":
                    texts.append(content.get("text", ""))

        return "\n".join(texts).strip()

    # ------------------------------------------------------------------
    # PROMPT
    # ------------------------------------------------------------------

    def _system_prompt(self) -> str:
        """Load system prompt for user stories generation."""
        return """
You are a Senior Business Analyst with deep BFSI experience.
You specialize in converting Business Requirements into detailed User Stories.
You respond ONLY in valid JSON following the provided schema.
Use BFSI / CBUAE terminology and follow Agile best practices.
"""

    # ------------------------------------------------------------------
    # PROMPT TEMPLATES
    # ------------------------------------------------------------------

    def _user_prompt_template(self) -> str:
        """Template for user stories generation prompt."""
        return """
Generate user stories based on the provided BRD content for version {version}.

CRITICAL RULES:
- The "description" fields in the JSON SCHEMA are your specific instructions for what to generate for that key.
- Do NOT include the "description" keys in your JSON output; only return the data keys and their generated values.
- No markdown, no explanations.
- Use BFSI / CBUAE terminology.
- Use clear, actionable language.
- Do NOT invent information not present in the BRD content.
- Format all text content using dot bullet points (starting with "• ") instead of paragraphs.

=== BRD CONTENT ===
{brd_content}

=== VERSION ===
{version}

=== JSON SCHEMA ===
{schema_json}
"""

    # ------------------------------------------------------------------
    # MODULE IDENTIFICATION
    # ------------------------------------------------------------------

    def identify_modules(self, brd_content: str) -> List[str]:
        """
        Identify high-level functional modules or epic-level areas in the BRD.
        This enables modular generation for better coverage.
        """
        prompt = f"""
Based on this BRD content, identify the high-level functional modules or areas.
Look for distinct business capabilities, user flows, or system components.
Return ONLY a JSON list of module names (3-8 modules total).

Examples:
- User Authentication & Security
- Payment Processing
- Admin Dashboard & Reporting
- Customer Onboarding
- Transaction Management

BRD Content:
{brd_content[:15000]}  # Use larger chunk for better module identification
"""
        
        response = self.client.responses.create(
            model=self.model,
            input=[
                {"role": "system", "content": "You are a Business Analyst identifying functional modules in BRD documents."},
                {"role": "user", "content": prompt}
            ],
            max_output_tokens=3000,
            temperature=0.1
        )
        
        # Extract text using the same method as generate_user_stories_json
        if hasattr(response, "output_text") and response.output_text:
            raw_response = response.output_text.strip()
        else:
            texts = []
            for item in getattr(response, "output", []):
                for content in getattr(item, "content", []):
                    if content.get("type") == "output_text":
                        texts.append(content.get("text", ""))
            raw_response = "\n".join(texts).strip()
        
        try:
            modules = json.loads(raw_response)
            print(f"DEBUG: Identified modules: {modules}")
            return modules
        except json.JSONDecodeError:
            # Fallback if JSON parsing fails
            print("DEBUG: Module identification failed, using fallback approach")
            return ["Core Functionality", "User Management", "Reporting & Analytics"]

    # ------------------------------------------------------------------
    # RAG PIPELINE FOR BRD EMBEDDINGS
    # ------------------------------------------------------------------

    def create_brd_embeddings(self, brd_file_path: str, embeddings_path: str) -> bool:
        """
        Create embeddings for a BRD document using LlamaIndex.
        This enables RAG-based retrieval for better User Stories generation.
        """
        try:
            print(f"DEBUG: Creating embeddings for BRD: {brd_file_path}")
            
            # Use LlamaParse to extract content from BRD document
            llama_parser = LlamaParse(result_type="markdown", api_key=os.getenv("LLAMA_CLOUD_API_KEY"))
            documents = llama_parser.load_data(brd_file_path)
            print(f"DEBUG: Parsed {len(documents)} documents from BRD")

            # Use MarkdownElementNodeParser for better multimodal support
            node_parser = MarkdownElementNodeParser(llm=self.llm, num_workers=4)
            nodes = node_parser.get_nodes_from_documents(documents)
            base_nodes, objects = node_parser.get_nodes_and_objects(nodes)
            print(f"DEBUG: Created {len(base_nodes)} base nodes and {len(objects)} objects")

            # Create VectorStoreIndex with multimodal support
            index = VectorStoreIndex(base_nodes + objects, embed_model=self.embed_model)
            print(f"DEBUG: Created VectorStoreIndex")

            # Persist embeddings
            index.storage_context.persist(embeddings_path)
            print(f"DEBUG: Persisted BRD embeddings to {embeddings_path}")
            
            return True
            
        except Exception as e:
            print(f"ERROR: Failed to create BRD embeddings: {e}")
            import traceback
            traceback.print_exc()
            return False

    def load_index(self, embeddings_path: str):
        """Loads and caches the index once for optimization."""
        if not embeddings_path or not os.path.exists(embeddings_path):
            print(f"ERROR: Embeddings path {embeddings_path} invalid or not found.")
            return None
        
        try:
            # Clear previous index to ensure clean loading
            self.index = None
            self.current_embeddings_path = None
            
            storage_context = StorageContext.from_defaults(persist_dir=embeddings_path)
            self.index = load_index_from_storage(storage_context)
            self.current_embeddings_path = embeddings_path
            print(f"DEBUG: Successfully loaded index from {embeddings_path}")
            return self.index
        except Exception as e:
            print(f"ERROR: Failed to load index: {e}")
            return None

    def retrieve_brd_context(self, embeddings_path: str, query: str, top_k: int = 20) -> str:
        """
        Enhanced retrieval with a global context anchor and resilient fallback.
        Uses persistent index loading for better performance.
        """
        # Debug: Check if embeddings_path is None
        if embeddings_path is None:
            print(f"DEBUG: retrieve_brd_context called with embeddings_path=None")
            return ""
        
        # Check if we need to load a new index
        if not self.index or self.current_embeddings_path != embeddings_path:
            if not self.load_index(embeddings_path):
                return ""
        
        try:
            # Use cached index for retrieval
            retriever = self.index.as_retriever(similarity_top_k=top_k)

            # Retrieve global overview AND module-specific nodes
            global_nodes = retriever.retrieve("Project Overview and High Level Requirements")
            local_nodes = retriever.retrieve(query)
            
            # Debug logging for retrieval quality
            if not global_nodes and not local_nodes:
                print(f"DEBUG: Retriever returned 0 nodes for query: {query}")
                return ""
            
            context_parts = []
            seen_text = set()
            
            # Use NodeWithScore correctly
            for node_with_score in (global_nodes[:5] + local_nodes):
                # IMPORTANT: Access the actual 'node' object inside the wrapper
                node = node_with_score.node 
                
                # Log similarity score for debugging
                print(f"DEBUG: Node similarity score: {node_with_score.score}")
                
                # Now extract the content from the actual node
                txt = node.get_content()
                
                if txt and txt not in seen_text:
                    context_parts.append(txt)
                    seen_text.add(txt)

            context = "\n\n".join(context_parts)
            print(f"DEBUG: Retrieved BRD context length: {len(context)} characters")
            
            return context
            
        except Exception as e:
            print(f"DEBUG: RAG Retrieval error: {e}")
            return ""

    # ------------------------------------------------------------------
    # JSON GENERATION
    # ------------------------------------------------------------------

    def generate_user_stories_json(self, brd_content: str, version: str, schema_json: str) -> Dict:
        user_prompt = self._user_prompt_template().format(
            brd_content=brd_content,
            version=version,
            schema_json=schema_json
        )

        response = self.client.responses.create(
            model=self.model,
            input=[
                {"role": "system", "content": self._system_prompt()},
                {"role": "user", "content": user_prompt}
            ],
            max_output_tokens=self.max_output_tokens,
            timeout=self.timeout
        )

        raw = self._extract_text(response)

        if not raw:
            raise RuntimeError("GPT returned empty output")

        print(f"DEBUG: Raw LLM response preview: {raw[:500]}...")

        try:
            data = json.loads(raw)
        except json.JSONDecodeError as e:
            raise RuntimeError(f"Invalid JSON from GPT: {e}\n\n{raw}")

        return data

    # ------------------------------------------------------------------
    # EXCEL EXPORT
    # ------------------------------------------------------------------

    def export_to_excel(self, user_stories_data: Dict, output_path: str) -> str:
        """Export user stories to Excel with multiple sheets for stories, epics, and dependencies."""
        try:
            # Use a context manager to ensure the file is always closed properly
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # --- SHEET 1: USER STORIES ---
                if 'user_stories' in user_stories_data:
                    expected_columns = [
                        'story_id', 'title', 'user_role', 'description', 
                        'acceptance_criteria', 'priority', 'effort_estimate', 
                        'brd_reference', 'version'
                    ]
                    
                    clean_stories_data = []
                    for story in user_stories_data['user_stories']:
                        clean_story = {}
                        for col in expected_columns:
                            value = story.get(col) or story.get(col.upper()) or story.get(col.title()) or story.get(col.capitalize())
                            if value and isinstance(value, str):
                                value = value.replace('•', '').strip()
                            clean_story[col] = value if value is not None else ""
                        clean_stories_data.append(clean_story)
                    
                    stories_df = pd.DataFrame(clean_stories_data)
                    stories_df.to_excel(writer, sheet_name='User Stories', index=False)
                    
                    # FORMATTING: Use column letters (A, B, C...) instead of names
                    worksheet = writer.sheets['User Stories']
                    column_widths = [15, 40, 25, 60, 60, 15, 20, 25, 15] # Widths matching expected_columns order
                    
                    for i, width in enumerate(column_widths):
                        col_letter = get_column_letter(i + 1)
                        worksheet.column_dimensions[col_letter].width = width

                # --- SHEET 2: EPICS ---
                if 'epics' in user_stories_data:
                    epics_df = pd.DataFrame(user_stories_data['epics'])
                    # Convert list values (like related_stories) to strings to prevent Excel errors
                    for col in epics_df.columns:
                        epics_df[col] = epics_df[col].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
                    
                    epics_df.to_excel(writer, sheet_name='Epics', index=False)
                    
                    # Simple auto-width for Epics
                    worksheet = writer.sheets['Epics']
                    for i in range(len(epics_df.columns)):
                        worksheet.column_dimensions[get_column_letter(i + 1)].width = 30

                # --- SHEET 3: DEPENDENCIES ---
                if 'dependencies' in user_stories_data:
                    deps_df = pd.DataFrame(user_stories_data['dependencies'])
                    deps_df.to_excel(writer, sheet_name='Dependencies', index=False)
                    
            print(f"DEBUG: Excel successfully saved to {output_path}")
            return output_path
            
        except Exception as e:
            print(f"ERROR: Failed to export to Excel: {e}")
            import traceback
            traceback.print_exc()
            # Ensure the potentially corrupt file is removed if it failed
            if os.path.exists(output_path):
                try: os.remove(output_path)
                except: pass
            raise RuntimeError(f"Excel export failed: {e}")

    # ------------------------------------------------------------------
    # MODULAR GENERATION APPROACH
    # ------------------------------------------------------------------

    def generate_extensive_user_stories(
        self,
        brd_content: str,
        version: str,
        schema_json: str,
        embeddings_path: str = None
    ) -> Dict:
        """
        Generate extensive user stories using resilient retrieval strategy.
        This method breaks down the BRD into modules and generates stories for each module.
        """
        print("DEBUG: Starting modular user stories generation...")
        
        # Step 1: Identify modules in the BRD
        modules = self.identify_modules(brd_content)
        
        all_stories_list = []
        all_epics = []
        all_deps = []
        
        # Global counters strictly managed outside the LLM call
        story_counter = 1
        epic_counter = 1

        # Create a "Foundation Context" (First 12k chars) to use if RAG fails
        foundation_context = brd_content[:12000]

        for i, module_name in enumerate(modules):
            print(f"DEBUG: Processing module {i+1}/{len(modules)}: {module_name}")
            
            # 1. Try RAG with multi-step retrieval for better context
            retrieval_query = f"{module_name} requirements, functional specifications, and business rules"
            
            # Retrieve technical constraints and data specifications
            tech_context = self.retrieve_brd_context(embeddings_path, "Technical constraints and data specifications", top_k=5)
            
            # Retrieve module-specific context
            module_context = self.retrieve_brd_context(embeddings_path, retrieval_query, top_k=15)
            
            # Combine contexts for richer information
            combined_context = f"{tech_context}\n\n{module_context}"
            
            # Debug: Check if embeddings_path is None
            if embeddings_path is None:
                print(f"DEBUG: embeddings_path is None for module {module_name}")

            # 2. SAFETY FALLBACK: If RAG returns < 500 chars, it likely failed. 
            # Use the module name + Foundation Context instead.
            if len(combined_context) < 500:
                print(f"INFO: RAG thin for {module_name} ({len(combined_context)} chars). Using Foundation Fallback.")
                combined_context = f"MODULE: {module_name}\n\nGENERAL BRD CONTEXT:\n{foundation_context}"

            # 3. Generate
            try:
                user_prompt = self._user_prompt_template().format(
                    brd_content=combined_context,
                    version=version,
                    schema_json=schema_json
                )

                response = self.client.responses.create(
                    model=self.model,
                    input=[
                        {"role": "system", "content": self._system_prompt()},
                        {"role": "user", "content": user_prompt}
                    ],
                    max_output_tokens=self.max_output_tokens
                )
                
                raw_json = self._extract_text(response)
                module_data = json.loads(raw_json)
                
                stories = module_data.get("user_stories", [])
                # Create a mapping from module story IDs to global story IDs
                story_id_mapping = {}
                
                for s in stories:
                    old_story_id = s.get("story_id", f"US-{story_counter:03d}")
                    new_story_id = f"US-{story_counter:03d}"
                    s["story_id"] = new_story_id
                    s["version"] = version
                    story_counter += 1
                    all_stories_list.append(s)
                    # Store mapping for epics and dependencies
                    story_id_mapping[old_story_id] = new_story_id

                # Process epics and update story IDs
                module_epics = module_data.get("epics", [])
                for epic in module_epics:
                    # Update epic_id to be globally unique
                    old_epic_id = epic.get("epic_id", f"EP-{epic_counter:03d}")
                    new_epic_id = f"EP-{epic_counter:03d}"
                    epic["epic_id"] = new_epic_id
                    epic_counter += 1
                    
                    if "related_stories" in epic and isinstance(epic["related_stories"], list):
                        # Map module story IDs to global story IDs
                        mapped_stories = []
                        for story_id in epic["related_stories"]:
                            if story_id in story_id_mapping:
                                mapped_stories.append(story_id_mapping[story_id])
                            else:
                                # If mapping not found, try to create a global ID
                                # This handles cases where story IDs might be in different formats
                                mapped_stories.append(story_id)
                        epic["related_stories"] = mapped_stories
                    all_epics.append(epic)

                # Process dependencies and update story IDs
                module_deps = module_data.get("dependencies", [])
                for dep in module_deps:
                    # Update story_id and depends_on fields
                    if "story_id" in dep:
                        old_story_id = dep["story_id"]
                        if old_story_id in story_id_mapping:
                            dep["story_id"] = story_id_mapping[old_story_id]
                    
                    if "depends_on" in dep:
                        old_depends_on = dep["depends_on"]
                        if old_depends_on in story_id_mapping:
                            dep["depends_on"] = story_id_mapping[old_depends_on]
                    
                    all_deps.append(dep)
                
                print(f"DEBUG: Added {len(stories)} stories for {module_name}")

            except Exception as e:
                print(f"ERROR: Module {module_name} failed: {e}")

        # Final aggregation
        return {
            "user_stories": all_stories_list,
            "epics": all_epics,
            "dependencies": all_deps
        }

    # ------------------------------------------------------------------
    # ORCHESTRATOR
    # ------------------------------------------------------------------

    def generate_user_stories_excel(
        self,
        brd_content: str,
        version: str,
        schema_json: str,
        output_path: str,
        use_modular_approach: bool = True,
        embeddings_path: str = None
    ) -> str:
        """Generate user stories from BRD content and export to Excel."""
        if use_modular_approach:
            data = self.generate_extensive_user_stories(brd_content, version, schema_json, embeddings_path)
        else:
            data = self.generate_user_stories_json(brd_content, version, schema_json)
        
        return self.export_to_excel(data, output_path)
